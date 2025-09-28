import openpyxl

class ExcelHandler:
    @staticmethod
    def read_excel(file_path, sheet_name=None):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name] if sheet_name else wb.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        wb.close()
        return data

    @staticmethod
    def write_excel(file_path, data, sheet_name=None):
        wb = openpyxl.Workbook()
        sheet = wb.active
        if sheet_name:
            sheet.title = sheet_name
        for row in data:
            sheet.append(row)
        wb.save(file_path)
        wb.close()
